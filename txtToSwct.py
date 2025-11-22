
from openpyxl import load_workbook
from pathlib import Path
import re
import math
from openpyxl.utils import get_column_letter, column_index_from_string
from Main import savePath
from Main import SWCTname


# --- настройки ---

XL_PATH_IN = Path("SWCTmacross.xlsm")
XL_PATH_OUT = Path(f"{SWCTname}.xlsm")  # сохраняем как .xlsm, чтобы не потерять VBA
SHEET_IDX   = 0
START_ROW   = 9        # начинаем всегда с 9-й строки
SHIFT_COLS  = 8        # сдвиг вправо на 7 колонок (B/C/D -> I/J/K)

PathList = []


def to_float(s: str):
    """Пробуем преобразовать в float и округлить вверх до целого."""
    if s is None:
        return None
    s = str(s).strip().replace(",", ".")
    try:
        value = float(s)
        return math.ceil(value)   # округляем вверх
    except ValueError:
        return None

def shift_col(col_letter: str, shift: int):
    idx = column_index_from_string(col_letter)
    return get_column_letter(idx + shift)

def collect_numbered_txt_files():
    """
    Ищем *.txt, имя начинается с: <число>.<пробел>...
    Примеры: '1. какой то текст.txt', '2. заголовок.txt'
    """
    out = []
    rx = re.compile(r'^(\d+)\.\s*(.+)\.txt$', re.IGNORECASE)


    ################################### PROBLEM!!!!!!!!!!!
    for p in PathList:
        if not Path(p).is_file() or Path(p).suffix.lower() != ".txt":
            continue
        m = rx.match(Path(p).name)
        if m:
            num = int(m.group(1))
            out.append((num, Path(p)))
    # сортируем по числовому префиксу
    p = Path(p)
    return [p for _, p in sorted(out, key=lambda t: t[0])]

def first_empty_row(ws, col_letter: str, start_row: int) -> int:
    """Ищем первую пустую строку в колонке col_letter начиная с start_row."""
    row = start_row
    while ws[f"{col_letter}{row}"].value not in (None, ""):
        row += 1
    return row

def main():

    
    if XL_PATH_OUT.suffix.lower() != ".xlsm":
        raise ValueError("Для сохранения макросов укажи выходной файл с расширением .xlsm.")

    if not XL_PATH_IN.exists():
        raise FileNotFoundError(f"Не найден входной файл: {XL_PATH_IN}")

    # ВАЖНО: keep_vba=True, чтобы сохранить встроенный проект VBA
    wb = load_workbook(XL_PATH_IN, keep_vba=True)
    ws = wb[wb.sheetnames[SHEET_IDX]]

    # Базовые колонки до сдвига
    TEXT_COL_BASE = "B"
    NUM1_COL_BASE = "C"
    NUM2_COL_BASE = "D"

    # Сдвинутые колонки (будут I/J/K)
    TEXT_COL = shift_col(TEXT_COL_BASE, SHIFT_COLS)
    NUM1_COL = shift_col(NUM1_COL_BASE, SHIFT_COLS)
    NUM2_COL = shift_col(NUM2_COL_BASE, SHIFT_COLS)

    files = collect_numbered_txt_files()
    if not files:
        print("Не найдено файлов вида 'N. что-то.txt' в текущей папке.")
        return

    total = 0
    row_ptr = first_empty_row(ws, TEXT_COL, START_ROW)

    for fpath in files:
        rows = []
        with fpath.open(encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                parts = line.split("\t")
                if len(parts) >= 2:
                    text = parts[1].strip()
                    num1 = to_float(parts[0])
                    num2 = to_float(parts[2]) if len(parts) >= 3 else None
                    rows.append((text, num1, num2))

        for text, num1, num2 in rows:
            ws[f"{TEXT_COL}{row_ptr}"] = text
            ws[f"{NUM1_COL}{row_ptr}"] = num1
            if num2 is not None:
                ws[f"{NUM2_COL}{row_ptr}"] = num2
            row_ptr += 1

        print(f"{fpath.name}: добавлено строк — {len(rows)}")
        total += len(rows)

    wb.save(XL_PATH_OUT)
    print(f"{SWCTname} Готово. Всего добавлено строк: {total}. Сохранено в: {XL_PATH_OUT}")
