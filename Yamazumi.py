##########################
###SWCT TO YAMAZUMI CEX###
##########################

from openpyxl import load_workbook
from pathlib import Path

from PySide6.QtCore import (
    QObject,
    QRunnable,
    QThreadPool,
    QTimer,
    Signal,
    Slot,
)
class WorkerSignals(QObject):

    finished = Signal()
    


class Yamazumi(QRunnable):

    def __init__(self, filePath, savePath):
        super().__init__()
        self.filePath = filePath
        self.savePath = savePath
        self.signals = WorkerSignals()


    def createOperationsList(self, sheet):
        num = 9
        sitesList = []
        site = "Value"
        while num <= self.counter:
            cell_f = sheet[f"F{num}"]
            cell_e = sheet[f"E{num}"]

            if isinstance(cell_e.value, str) and isinstance(cell_f.value, str):
                site = sheet[f"E{num}"].value
                benefit = sheet[f"H{num}"].value
                loss = sheet[f"I{num}"].value
                # удаление всех технологических процессов
                if loss > 7200:
                    loss = loss - 7200
                if loss > 25000:
                    loss = 0
                sitesList.append({"Операция": sheet[f"F{num}"].value, "Участок": site, "Польза": benefit,"Потери": loss})

            elif cell_e.value is None and isinstance(cell_f.value, str):
                # удаление всех технологических процессов
                loss = sheet[f"I{num}"].value
                if loss > 7200:
                    loss = loss - 7200
                if loss > 15000:
                    loss = 0
                sitesList.append({"Операция":sheet[f"F{num}"].value, "Участок": site,"Польза": sheet[f"H{num}"].value, "Потери": loss})
            num += 1
        return sitesList

    def writeInWorkshop(self, operationsList, sheet):
        siteCol = {"МЕХ СБ": [9,10,11], "ППМ": [14,15,16], "РЕГ":[19,20,21], "ПРСБ":[24,25,26], "ПРОГОН":[29,30,31], "ЛИНЗЫ":[34,35,36],"СБ":[39,40,41], "УП":[44,45,46]}
        number = 16
        site = 0
        for i in operationsList:
            if site == 0:
                site = i.get('Участок')

            if site != i.get('Участок'):
                number = 16
                site = i.get('Участок')

            if siteCol.get(i.get('Участок')) != None and sheet.cell(row = number, column = siteCol[i.get('Участок')][1]).value == None:
                sheet.cell(row = number, column = siteCol[i.get('Участок')][0], value = i.get('Операция'))
                sheet.cell(row = number, column = siteCol[i.get('Участок')][1], value = i.get('Польза'))
                sheet.cell(row = number, column = siteCol[i.get('Участок')][2], value = i.get('Потери'))
                number += 1

            else:
                while sheet.cell(row = number, column = siteCol[i.get('Участок')][1]).value != None:
                    number += 1

                sheet.cell(row = number, column = siteCol[i.get('Участок')][0], value = i.get('Операция'))
                sheet.cell(row = number, column = siteCol[i.get('Участок')][1], value = i.get('Польза'))
                sheet.cell(row = number, column = siteCol[i.get('Участок')][2], value = i.get('Потери'))
                number += 1
    
    @Slot()    
    def run(self):

        templateWorkshop = Path("Yamazumi.xlsx")
        templateWorkshop = load_workbook(templateWorkshop)

        activeWb = load_workbook(Path(self.filePath))
        sheet = activeWb["SWCT"]

        self.counter = 9
        for row in sheet["J9:J300"]:
            for cell in row:
                if cell.value != None:
                    self.counter += 1

        operationsList = self.createOperationsList(sheet)

        sheet = templateWorkshop["YAMAZUMI цеха"]

        self.writeInWorkshop(operationsList, sheet)

        print(f"{self.savePath}/Yamazumi цеха.xlsx")
        self.signals.finished.emit()
        templateWorkshop.save(f"{self.savePath}\Yamazumi цеха.xlsx")

if __name__ == "__main__":
    result = Yamazumi(r"C:\Users\pisos\Downloads", "SWCT Светильник LINE ILF30-1,5W40-30H-150(50P02)1.xlsm")
    QThreadPool.globalInstance().start(result)
    result.run()