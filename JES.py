###JES CREATOR###
from openpyxl.styles import Border, Side
from openpyxl import load_workbook
from pathlib import Path
from functools import reduce
import os
from datetime import date

# xlPath = Path("/content/drive/MyDrive/Yamazumi/$SWCT Светильник LINE ILF30-1,5W40-30H-150(50P02).xlsm")
# pathTemplateJES = Path("/content/drive/MyDrive/Yamazumi/JES.xlsx")
# copyTemplateJES = Path("CopyJES.xlsx")

# activeWb = load_workbook(xlPath, keep_vba=True)
# templateJES = load_workbook(pathTemplateJES)

class Jes():
    def __init__(self, xlPath, savePath):
        self.savePath = savePath
        self.pathTemplateJES = Path("JES.xlsx")
        self.templateJES = load_workbook(self.pathTemplateJES)
        self.xlPath = Path(xlPath)
        self.activeWb = load_workbook(self.xlPath)
        # self.copyTemplateJES = load_workbook(Path("CopyJES.xlsx"))
        pass

    def count_files_recursive(self, folder_path):
        path = Path(folder_path)
        # rglob('*') ищет во всех подпапках
        # is_file() проверяет, что это файл
        file_count = sum(1 for file in path.rglob('*') if file.is_file())
        return file_count


    def number_within_group(self, group):
        group = group.copy()
        group['group_number'] = range(1, len(group) + 1)
        return group

    def getLampName(self):
        lampsList = ["BOX", "DOT", "EDGE", "GROUND-mini", "IntBEAM", "IntDOT", "IntGROUND", "IntGROUND-midi", "IntLINE", "IntPOINT", "IntSLIM", "IntSLOT", "IntSPOT",
                    "IntSTARK", "IntTOP", "IntTUBE", "IntTWIN", "KUB", "LINE", "LINE-GROUND", "LINEUP", "MARK", "PILL", "RAY", "SLIM", "SPOT", "TWIN", "WALL", "ZENITH",
                    "Axceccyapы", "Допоборудование", "Модули", "ТУ", "Устройства управления"]
        for lamp in lampsList:
            if lamp in xlPath.stem:
                return lamp

    def getFullLampName(self):
        fullName = xlPath.stem.replace("$SWCT","").replace("Светильник","").rstrip()
        return fullName

    def createOperationsList(self, counter, sheet):
            
        num = 9
        sitesList = []
        site = "Value"
        Operation = "Value"
        OpBenefit = "Value"
        OpLoss = "Value"

        while num <= counter:
            cell_f = sheet[f"F{num}"]
            cell_m = sheet[f"M{num}"]

            #первое вхождение в участов
            if isinstance(cell_f.value, str)  and isinstance(cell_m.value, str):
                # counter = 1
                if sheet[f"K{num}"].value == None:
                    sheet[f"K{num}"].value = 0
                elif sheet[f"L{num}"].value == None:
                    sheet[f"L{num}"].value = 0

                site = sheet[f"M{num}"].value
                Operation = str(sheet[f"F{num}"].value)
                Operation = Operation.rstrip()
                OpBenefit = sheet[f"H{num}"].value
                OpLoss = sheet[f"I{num}"].value
                Step = str(sheet[f"J{num}"].value)
                Step = Step.rstrip()
                StepBenefit = sheet[f"K{num}"].value
                StepLoss = sheet[f"L{num}"].value



                sitesList.append({
                "Операция": {
                    "Название операции": Operation,
                    "Польза": OpBenefit,
                    "Потери": OpLoss,
                },
                "Шаг": {
                    "Шаг операции": Step,
                    "Польза шага": StepBenefit,
                    "Потеря шага": StepLoss
                },
                "Участок": site
            })


            #повторное вхожднние в участок
            elif cell_m.value is None and cell_f.value is None:

                        if sheet[f"K{num}"].value == None:
                            sheet[f"K{num}"].value = 0
                        elif sheet[f"L{num}"].value == None:
                            sheet[f"L{num}"].value = 0

                        sitesList.append({
                "Операция": {
                    "Название операции": Operation,
                    "Польза": OpBenefit,
                    "Потери": OpLoss,
                },
                "Шаг": {
                    "Шаг операции": sheet[f"J{num}"].value,
                    "Польза шага": sheet[f"K{num}"].value,
                    "Потеря шага": sheet[f"L{num}"].value
                },
                "Участок": site
            })

            elif isinstance(cell_f.value, str)  and cell_m.value is None:

                        if sheet[f"K{num}"].value == None:
                            sheet[f"K{num}"].value = 0
                        elif sheet[f"L{num}"].value == None:
                            sheet[f"L{num}"].value = 0

                        Operation = sheet[f"F{num}"].value
                        Operation = Operation.rstrip()
                        OpBenefit = sheet[f"H{num}"].value
                        OpLoss = sheet[f"I{num}"].value

                        sitesList.append({
                "Операция": {
                    "Название операции": str(sheet[f"F{num}"].value).rstrip(),
                    "Польза": sheet[f"H{num}"].value,
                    "Потери": sheet[f"I{num}"].value,
                },
                "Шаг": {
                    "Шаг операции": sheet[f"J{num}"].value,
                    "Польза шага": sheet[f"K{num}"].value,
                    "Потеря шага": sheet[f"L{num}"].value
                },
                "Участок": site
            })


            num += 1

        return sitesList

    def deleteEmptyPages(self, workBook):
        sheets_to_delete = []

        for page in range(1, 6):
            sheet_name = f"{page}"
            if sheet_name in workBook.sheetnames:
                sheet = workBook[sheet_name]
                if sheet["B8"].value is None:
                    sheets_to_delete.append(sheet_name)

        # Удаляем листы после цикла (чтобы не нарушать итерацию)
        for sheet_name in sheets_to_delete:
            del workBook[sheet_name]

        return workBook

    def addList(self, WorkBook, activeList):
        sheet = WorkBook[f"{activeList}"]
        return sheet

    def copy_workbook(self,source_path, destination_path):
        # Загружаем исходную книгу
        source_wb = load_workbook(source_path)

        # Создаем новую книгу или сохраняем под новым именем
        source_wb.save(destination_path)


    def creatFolders(self, templateJES, copyTemplateJES, pathTemplateJES):
        # currentData = str(date.today().strftime("%d/%m/%Y"))
        sheet = templateJES["1"]
        name = "value"
        OpName = "value"
        self.jesPath = f"{self.savePath}/JES"
        os.makedirs(self.jesPath, exist_ok = True)
        num = 8
        timer = 0

        #создаем папки участков
        for i in self.operationsList:
            #создаем папку
            try:
                if i.get("Участок") != name:
                    name = i.get("Участок")
                    os.makedirs(f"{self.jesPath}/JES {name}")
            except:
                pass


            #первое вхождение в операцию
            if i.get("Операция").get("Название операции") != OpName:

                #Великое сохранение файла
                if timer != 0:
                    print(OpName)
                    wb = self.deleteEmptyPages(wb)
                    count = self.count_files_recursive(f"/content/drive/MyDrive/JES/JES {site}/")
                    wb.save(f"/content/drive/MyDrive/JES/JES {site}/{count +1 }. {OpName.replace("/", "")}.xlsx")

                self.copy_workbook(pathTemplateJES, copyTemplateJES) #создаем копию JES
                wb = load_workbook(copyTemplateJES)#открываем JES
                sheet = wb["1"]
                # Откатили num шагов
                num = 8
                activeList = 1
                OpName = i.get("Операция").get("Название операции") # Запись операции
                TactTime = int(i.get("Операция").get("Польза")) + int(i.get("Операция").get("Потери"))
                Time = int(i.get("Шаг").get("Польза шага")) + int(i.get("Шаг").get("Потеря шага")) # Запись Т
                site = i.get("Участок")

                sheet["I2"] = self.getFullLampName()
                sheet["U2"] = f"Изготовление светильника {self.getLampName()}"
                # sheet["AT2"] = currentData

                sheet["BF2"] = activeList
                sheet["AP2"] = TactTime
                sheet["A5"] = OpName
                sheet[f"B{num}"] = i.get("Шаг").get("Шаг операции") # Запись первого шага
                sheet[f"BG{num}"] = Time
                timer += 1

            #повторное вхождение в операцию
            elif i.get("Операция").get("Название операции") == OpName:

                num += 1
                #добавление листа
                if num == 12:
                    activeList += 1
                    sheet = self.addList(wb, activeList)
                    sheet["A5"] = OpName
                    sheet["BF2"] = activeList
                    sheet["AP2"] = TactTime
                    sheet["I2"] = self.getFullLampName()
                    sheet["U2"] = f"Изготовление светильника {self.getLampName()}"
                    # sheet["AT2"] = currentData
                    num = 8


                sheet[f"B{num}"] = i.get("Шаг").get("Шаг операции") # записали шаг

                step_data = i.get("Шаг", {})
                benefit = step_data.get("Польза шага", 0)
                loss = step_data.get("Потеря шага", 0)

                try:
                    # Пробуем преобразовать оба значения в числа
                    Time = int(benefit) + int(loss)
                except (ValueError, TypeError):
                    try:
                        # Если не получилось, пробуем вычислить выражение
                        has_letters = any(char.isalpha() for char in str(loss)) #### Удалить если не работает
                        if has_letters == True:
                            loss = 0

                        Time = int(benefit) + eval(str(loss))
                    except:
                        # Если всё равно ошибка, используем значения по умолчанию
                        Time = 0

                sheet[f"BG{num}"] = Time


    def main(self):

        sheet = self.activeWb["SWCT"]

        counter = 9
        for row in sheet["J9:J300"]:
            for cell in row:
                if cell.value != None:
                    counter += 1


        self.operationsList = self.createOperationsList(counter, sheet)

        # for i in operationsList:
        #     print(i)

        # with open("List.txt", "w", encoding="utf-8") as file:
        #     for i in operationsList:
        #         file.write(f"{i}\n")

        self.creatFolders(self.templateJES, self.copyTemplateJES, self.pathTemplateJES)



result = Jes(r"C:\Users\pisos\Downloads\SWCT LINE ILF30-1,5W40-30H-150(50P02)1.xlsm")
result.main()

