
from openpyxl import load_workbook
from pathlib import Path
import re
import math
from openpyxl.utils import get_column_letter, column_index_from_string

from PyQt5.QtCore import (
    QObject,
    QRunnable,
    QThreadPool,
    QTimer,
    pyqtSignal,
    pyqtSlot,
)
# import asyncio
# from Main import savePath
# from Main import SWCTname


# XL_PATH_IN = Path("SWCTmacross.xlsm")
# XL_PATH_OUT = Path(".xlsm")  # сохраняем как .xlsm, чтобы не потерять VBA
# SHEET_IDX   = 0
# START_ROW   = 9        # начинаем всегда с 9-й строки
# SHIFT_COLS  = 8        # сдвиг вправо на 7 колонок (B/C/D -> I/J/K)
# PathList = []

class WorkerSignals(QObject):
    """Signals from a running worker thread.

    finished
        No data

    error
        tuple (exctype, value, traceback.format_exc())

    result
        object data returned from processing, anything

    progress
        float indicating % progress
    """

    finished = pyqtSignal()
    error = pyqtSignal(tuple)
    result = pyqtSignal(str)
    progress = pyqtSignal(float)


class Text(QRunnable):

    def __init__(self, pathList, SWCTname, savePath):
        super().__init__()
        self.pathList = pathList
        self.SWCTname = SWCTname
        self.savePath = savePath

    def to_float(self, s: str):

        if s is None:
            return None
        self.s = str(s).strip().replace(",", ".")
        try:
            value = float(s)
            return math.ceil(value)   # округляем вверх
        except ValueError:
            return None
        
    def shift_col(self, col_letter: str, shift: int):
        self.idx = column_index_from_string(col_letter)
        return get_column_letter(self.idx + shift)
    
    def collect_numbered_txt_files(self, pathList):

        """
        Ищем *.txt, имя начинается с: <число>.<пробел>...
        Примеры: '1. какой то текст.txt', '2. заголовок.txt'
        """
        self.out = []
        self.rx = re.compile(r'^(\d+)\.\s*(.+)\.txt$', re.IGNORECASE)
        ################################### PROBLEM!!!!!!!!!!!
        for p in pathList:
            if not Path(p).is_file() or Path(p).suffix.lower() != ".txt":
                continue
            self.m = self.rx.match(Path(p).name)
            if self.m:
                self.num = int(self.m.group(1))
                self.out.append((self.num, Path(p)))
        # сортируем по числовому префиксу
        p = Path(p)


        return [p for _, p in sorted(self.out, key=lambda t: t[0])]
    
    def first_empty_row(self, ws, col_letter: str) -> int:
        """Ищем первую пустую строку в колонке col_letter начиная с start_row."""
        self.row = 9
        while ws[f"{col_letter}{self.row}"].value not in (None, ""):
            self.row += 1
        return self.row
    

    @pyqtSlot()
    def run(self):

        XL_PATH_IN = "SWCTmacross.xlsm"

        if Path(f"{self.savePath}/{self.SWCTname}").is_file():
            XL_PATH_IN = f"{self.savePath}/{self.SWCTname}"

        self.wb = load_workbook(XL_PATH_IN, keep_vba=True)
        ws = self.wb.active

        # Сдвинутые колонки (будут I/J/K)
        TEXT_COL = "J"
        NUM1_COL = "K"
        NUM2_COL = "L"

        files = self.collect_numbered_txt_files(self.pathList)

        if not files:
            print("Не найдено файлов вида 'N. что-то.txt' в текущей папке.")
            return

        total = 0
        row_ptr = self.first_empty_row(ws, TEXT_COL)


        for fpath in files:
            rows = []
            with fpath.open(encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    parts = line.split("\t")
                    if len(parts) >= 2:
                        text = parts[1].strip()
                        num1 = self.to_float(parts[0])
                        num2 = self.to_float(parts[2]) if len(parts) >= 3 else None
                        rows.append((text, num1, num2))

            for text, num1, num2 in rows:
                ws[f"{TEXT_COL}{row_ptr}"] = text
                ws[f"{NUM1_COL}{row_ptr}"] = num1
                if num2 is not None:
                    ws[f"{NUM2_COL}{row_ptr}"] = num2
                row_ptr += 1



            #print(f"{fpath.name}: добавлено строк — {len(rows)}")
            total += len(rows)

        # self.signals = WorkerSignals()
        # self.signals.result.emit("result")
        self.wb.save(f"{self.savePath}\{self.SWCTname}")
        #print(f"Готово. Всего добавлено строк: {total}. Сохранено в: {self.savePath}/{self.SWCTname}")



